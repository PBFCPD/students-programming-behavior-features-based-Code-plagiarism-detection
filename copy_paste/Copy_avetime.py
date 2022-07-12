import openpyxl

cid = '2315'

data_path = r'../excel/7-' + cid + '/' + cid +'_time.xlsx'
workBook = openpyxl.load_workbook(data_path)
workSheet = workBook.worksheets[0]
saveFile = r'../excel/7-' + cid + '/' + cid +'_time.xlsx'

totaltime0 = 0
totaltime1 = 0
totaltime2 = 0
totaltime3 = 0
totaltime4 = 0
totaltime5 = 0
totaltime6 = 0
totaltime7 = 0
totaltime8 = 0
totaltime9 = 0
totaltime10 = 0
totaltime11 = 0
totaltime12 = 0
# totaltime13 = 0

i0 = 0
i1 = 0
i2 = 0
i3 = 0
i4 = 0
i5 = 0
i6 = 0
i7 = 0
i8 = 0
i9 = 0
i10 = 0
i11 = 0
i12 = 0
# i13 = 0
for row in range(2, workSheet.max_row+1):
    problem = workSheet.cell(row, 3).value
    ac = workSheet.cell(row, 4).value
    time = workSheet.cell(row, 5).value
    if(ac == 1):
        if(problem == 0):
            i0 = i0 + 1
            totaltime0 = totaltime0 + time
        elif (problem == 1):
            i1 = i1 + 1
            totaltime1 = totaltime1 + time
        elif (problem == 2):
            i2 = i2 + 1
            totaltime2 = totaltime2 + time
        elif (problem == 3):
            i3 = i3 + 1
            totaltime3 = totaltime3 + time
        elif (problem == 4):
            i4 = i4 + 1
            totaltime4 = totaltime4 + time
        elif (problem == 5):
            i5 = i5 + 1
            totaltime5 = totaltime5 + time
        elif (problem == 6):
            i6 = i6 + 1
            totaltime6 = totaltime6 + time
        elif (problem == 7):
            i7 = i7 + 1
            totaltime7 = totaltime7 + time
        elif (problem == 8):
            i8 = i8 + 1
            totaltime8 = totaltime8 + time
        elif (problem == 9):
            i9 = i9 + 1
            totaltime9 = totaltime9 + time
        elif (problem == 10):
            i10 = i10 + 1
            totaltime10 = totaltime10 + time
        elif (problem == 11):
            i11 = i11 + 1
            totaltime11 = totaltime11 + time
        elif (problem == 12):
            i12 = i12 + 1
            totaltime12 = totaltime12 + time
        # elif (problem == 13):
        #     i13 = i13 + 1
        #     totaltime13 = totaltime13 + time
    print(row)

avetime0 = round(totaltime0/i0)
avetime1 = round(totaltime1/i1)
avetime2 = round(totaltime2/i2)
avetime3 = round(totaltime3/i3)
avetime4 = round(totaltime4/i4)
avetime5 = round(totaltime5/i5)
avetime6 = round(totaltime6/i6)
avetime7 = round(totaltime7/i7)
avetime8 = round(totaltime8/i8)
avetime9 = round(totaltime9/i9)
avetime10 = round(totaltime10/i10)
avetime11 = round(totaltime11/i11)
avetime12 = round(totaltime12/i12)
# avetime13 = round(totaltime13/i13)
print(avetime0, avetime1, avetime2, avetime3, avetime4, avetime5)

for row in range(2, workSheet.max_row+1):
    problem = workSheet.cell(row, 3).value
    if(problem == 0):
        workSheet.cell(row, 6, avetime0)
    elif (problem == 1):
        workSheet.cell(row, 6, avetime1)
    elif (problem == 2):
        workSheet.cell(row, 6, avetime2)
    elif (problem == 3):
        workSheet.cell(row, 6, avetime3)
    elif (problem == 4):
        workSheet.cell(row, 6, avetime4)
    elif (problem == 5):
        workSheet.cell(row, 6, avetime5)
    elif (problem == 6):
        workSheet.cell(row, 6, avetime6)
    elif (problem == 7):
        workSheet.cell(row, 6, avetime7)
    elif (problem == 8):
        workSheet.cell(row, 6, avetime8)
    elif (problem == 9):
        workSheet.cell(row, 6, avetime9)
    elif (problem == 10):
        workSheet.cell(row, 6, avetime10)
    elif (problem == 11):
        workSheet.cell(row, 6, avetime11)
    elif (problem == 12):
        workSheet.cell(row, 6, avetime12)
    # elif (problem == 13):
    #     workSheet.cell(row, 6, avetime13)
    print(row)


    workBook.save(saveFile)


