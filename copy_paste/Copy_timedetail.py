import openpyxl
from pandas import DataFrame,read_excel

cid = '2315'

data_path = r'../excel/7-' + cid + '/' + cid +'_time.xlsx'
workBook = openpyxl.load_workbook(data_path)
workSheet = workBook.worksheets[0]
saveFile = r'../excel/7-' + cid + '/' + cid +'_time.xlsx'

workSheet.cell(1, 7, '花费时间 - 平均花费时间')
workSheet.cell(1, 8, 'ABS（花费时间-平均花费时间）')
workSheet.cell(1, 9, '平均花费时间 + 平均 * 0.2')
workSheet.cell(1, 10, '平均花费时间 - 平均 * 0.2')
workSheet.cell(1, 11, '超出0.2的部分')
workSheet.cell(1, 12, '平均花费时间 + 平均 * 0.3')
workSheet.cell(1, 13, '平均花费时间 - 平均 * 0.3')
workSheet.cell(1, 14, '超出0.3的部分')
workSheet.cell(1, 15, '平均花费时间 + 平均 * 0.4')
workSheet.cell(1, 16, '平均花费时间 - 平均 * 0.4')
workSheet.cell(1, 17, '超出0.4的部分')
workSheet.cell(1, 18, '平均花费时间 + 平均 * 0.5')
workSheet.cell(1, 19, '平均花费时间 - 平均 * 0.5')
workSheet.cell(1, 20, '超出0.5的部分')

workSheet.cell(1, 21, 'abs归一化')
workSheet.cell(1, 22, '0.2归一化')
workSheet.cell(1, 23, '0.3归一化')
workSheet.cell(1, 24, '0.4归一化')
workSheet.cell(1, 25, '0.5归一化')

for row in range(2, workSheet.max_row+1):
    time = workSheet.cell(row, 5).value
    avetime = workSheet.cell(row, 6).value

    # 花费时间 - 平均花费时间
    reducetime = time - avetime

    # ABS（花费时间-平均花费时间）
    absreducetime = abs(time - avetime)

    # 平均花费时间 + 平均 * 0.2
    avetime_12 = round(avetime + avetime * 0.2)

    # 平均花费时间 - 平均 * 0.2
    avetime_02 = round(avetime - avetime * 0.2)

    # 超出0.2的部分
    if(time > avetime_12):
        beyondtime_2 = time - avetime_12
    elif(time < avetime_02):
        beyondtime_2 = avetime_02 - time
    else:
        beyondtime_2 = 0

    # 平均花费时间 + 平均 * 0.3
    avetime_13 = round(avetime + avetime * 0.3)

    # 平均花费时间 - 平均 * 0.3
    avetime_03 = round(avetime - avetime * 0.3)

    # 超出0.3的部分
    if (time > avetime_13):
        beyondtime_3 = time - avetime_13
    elif (time < avetime_03):
        beyondtime_3 = avetime_03 - time
    else:
        beyondtime_3 = 0

    # 平均花费时间 + 平均 * 0.4
    avetime_14 = round(avetime + avetime * 0.4)

    # 平均花费时间 - 平均 * 0.4
    avetime_04 = round(avetime + avetime * 0.4)

    # 超出0.4的部分
    if (time > avetime_14):
        beyondtime_4 = time - avetime_14
    elif (time < avetime_04):
        beyondtime_4 = avetime_04 - time
    else:
        beyondtime_4 = 0

    # 平均花费时间 + 平均 * 0.5
    avetime_15 = round(avetime + avetime * 0.5)

    # 平均花费时间 - 平均 * 0.5
    avetime_05 = round(avetime + avetime * 0.5)

    # 超出0.5的部分
    if (time > avetime_15):
        beyondtime_5 = time - avetime_15
    elif (time < avetime_05):
        beyondtime_5 = avetime_05 - time
    else:
        beyondtime_5 = 0

    workSheet.cell(row, 7, reducetime)
    workSheet.cell(row, 8, absreducetime)
    workSheet.cell(row, 9, avetime_12)
    workSheet.cell(row, 10, avetime_02)
    workSheet.cell(row, 11, beyondtime_2)
    workSheet.cell(row, 12, avetime_13)
    workSheet.cell(row, 13, avetime_03)
    workSheet.cell(row, 14, beyondtime_3)
    workSheet.cell(row, 15, avetime_14)
    workSheet.cell(row, 16, avetime_04)
    workSheet.cell(row, 17, beyondtime_4)
    workSheet.cell(row, 18, avetime_15)
    workSheet.cell(row, 19, avetime_05)
    workSheet.cell(row, 20, beyondtime_5)
    workBook.save(saveFile)
    print(reducetime, absreducetime, avetime_12, avetime_02, beyondtime_2, avetime_13, avetime_03, beyondtime_3,avetime_14, avetime_04, beyondtime_4,avetime_15, avetime_05, beyondtime_5)

print("*******1 END*******")
data = DataFrame(read_excel(saveFile))

abs = list(data['ABS（花费时间-平均花费时间）'])
data2 = list(data['超出0.2的部分'])
data3 = list(data['超出0.3的部分'])
data4 = list(data['超出0.4的部分'])
data5 = list(data['超出0.5的部分'])

maxabs = max(abs)
maxdata2 = max(data2)
maxdata3 = max(data3)
maxdata4 = max(data4)
maxdata5 = max(data5)
print(maxabs, maxdata2, maxdata3, maxdata4, maxdata5)

for row in range(2, workSheet.max_row+1):
    absreducetime = workSheet.cell(row, 8).value
    beyondtime_2 = workSheet.cell(row, 11).value
    beyondtime_3 = workSheet.cell(row, 14).value
    beyondtime_4 = workSheet.cell(row, 17).value
    beyondtime_5 = workSheet.cell(row, 20).value

    workSheet.cell(row, 21, round(absreducetime/maxabs, 2))
    workSheet.cell(row, 22, round(beyondtime_2/maxdata2, 2))
    workSheet.cell(row, 23, round(beyondtime_3/maxdata3, 2))
    workSheet.cell(row, 24, round(beyondtime_4/maxdata4, 2))
    workSheet.cell(row, 25, round(beyondtime_5/maxdata5, 2))
    workBook.save(saveFile)
    print(round(absreducetime/maxabs, 2), round(beyondtime_2/maxdata2, 2), round(beyondtime_3/maxdata3, 2), round(beyondtime_4/maxdata4, 2), round(beyondtime_5/maxdata5, 2))



