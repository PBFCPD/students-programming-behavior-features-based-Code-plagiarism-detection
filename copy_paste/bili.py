# 按照比例计算 2262

# 1.最小化页面次数归一化  -  0.15
# 2.鼠标离开页面的时间归一化  -  0.15
# 3.异常粘贴次数  -  0.3
# 4.终端输入次数abs归一化	终端输入次数0.2归一化	终端输入次数0.3归一化	终端输入次数0.4归一化	终端输入次数0.5归一化  -  0.2
# 5.timeabs	time0.2	time0.3	time0.4	time0.5  -  0.2

import openpyxl
import pandas as pd
from decimal import Decimal
from pandas import DataFrame,read_excel

c_mini = 0.1
c_leave = 0.1
c_abnormal = 0.3
c_command = 0.2
c_time = 0.3

cid = '2262'
data_path = r'../excel/6-' + cid + '/' + cid +'_total_2.xlsx'
workBook = openpyxl.load_workbook(data_path)
workSheet = workBook.worksheets[1]
saveFile = r'../excel/6-' + cid + '/' + cid +'_total_2.xlsx'


workSheet.cell(1, 17, '最小化页面次数归一化')
workSheet.cell(1, 18, '鼠标离开页面的时间归一化')
workSheet.cell(1, 19, '异常粘贴次数')
workSheet.cell(1, 20, '终端输入次数abs归一化')
workSheet.cell(1, 21, '终端输入次数0.2归一化')
workSheet.cell(1, 22, '终端输入次数0.3归一化')
workSheet.cell(1, 23, '终端输入次数0.4归一化')
workSheet.cell(1, 24, '终端输入次数0.5归一化')
workSheet.cell(1, 25, 'timeabs')
workSheet.cell(1, 26, 'time0.2')
workSheet.cell(1, 27, 'time0.3')
workSheet.cell(1, 28, 'time0.4')
workSheet.cell(1, 29, 'time0.5')

data_path1 = r'../excel/6-' + cid + '/' + cid +'_total_2.xlsx'
data1 = DataFrame(read_excel(data_path1, sheet_name='Sheet2'))

datamini = data1.sort_values(by="time0.4", ascending=True)
username = list(datamini['用户名'])
minisize = list(datamini['time0.4'])
mini0 = 0
minidict = {}
for user, mini in zip(list(username), list(minisize)):
    if(mini == 0):
        mini0 = mini0 + 1
        miniper = 0
        minidict[user] = miniper
    else:
        mini0 = mini0 + 1
        # time
        if (mini0 >= 2 and mini0 <= 49):
            miniper = 0.06
            minidict[user] = miniper
        elif (mini0 >= 50 and mini0 <= 98):
            miniper = 0.12
            minidict[user] = miniper
        elif (mini0 >= 99 and mini0 <= 147):
            miniper = 0.18
            minidict[user] = miniper
        elif (mini0 >= 148 and mini0 <= 196):
            miniper = 0.24
            minidict[user] = miniper
        elif (mini0 >= 197 and mini0 <= 245):
            miniper = 0.3
            minidict[user] = miniper

        # 终端5
        # if (mini0 >= 8 and mini0 <= 86):
        #     miniper = 0.08
        #     minidict[user] = miniper
        # elif (mini0 >= 87 and mini0 <= 165):
        #     miniper = 0.13
        #     minidict[user] = miniper
        # elif (mini0 >= 166 and mini0 <= 245):
        #     miniper = 0.2
        #     minidict[user] = miniper

        # 终端4
        # if (mini0 >= 9 and mini0 <= 55):
        #     miniper = 0.04
        #     minidict[user] = miniper
        # elif (mini0 >= 56 and mini0 <= 102):
        #     miniper = 0.08
        #     minidict[user] = miniper
        # elif (mini0 >= 103 and mini0 <= 149):
        #     miniper = 0.12
        #     minidict[user] = miniper
        # elif (mini0 >= 150 and mini0 <= 197):
        #     miniper = 0.16
        #     minidict[user] = miniper
        # elif (mini0 >= 198 and mini0 <= 245):
        #     miniper = 0.2
        #     minidict[user] = miniper

        # 终端3
        # if (mini0 >= 43 and mini0 <= 109):
        #     miniper = 0.08
        #     minidict[user] = miniper
        # elif (mini0 >= 110 and mini0 <= 177):
        #     miniper = 0.13
        #     minidict[user] = miniper
        # elif (mini0 >= 178 and mini0 <= 245):
        #     miniper = 0.2
        #     minidict[user] = miniper

        # 终端2
        # if (mini0 >= 32 and mini0 <= 102):
        #     miniper = 0.08
        #     minidict[user] = miniper
        # elif (mini0 >= 103 and mini0 <= 173):
        #     miniper = 0.13
        #     minidict[user] = miniper
        # elif (mini0 >= 174 and mini0 <= 245):
        #     miniper = 0.2
        #     minidict[user] = miniper

        # 终端abs
        # if (mini0 >= 16 and mini0 <= 61):
        #     miniper = 0.05
        #     minidict[user] = miniper
        # elif (mini0 >= 62 and mini0 <= 107):
        #     miniper = 0.08
        #     minidict[user] = miniper
        # elif (mini0 >= 108 and mini0 <= 153):
        #     miniper = 0.12
        #     minidict[user] = miniper
        # elif (mini0 >= 154 and mini0 <= 199):
        #     miniper = 0.17
        #     minidict[user] = miniper
        # elif (mini0 >= 200 and mini0 <= 245):
        #     miniper = 0.2
        #     minidict[user] = miniper

        # 异常粘贴
        # if (mini0 >= 63 and mini0 <= 98):
        #     miniper = 0.06
        #     minidict[user] = miniper
        # elif (mini0 >= 99 and mini0 <= 134):
        #     miniper = 0.12
        #     minidict[user] = miniper
        # elif (mini0 >= 135 and mini0 <= 171):
        #     miniper = 0.18
        #     minidict[user] = miniper
        # elif (mini0 >= 172 and mini0 <= 208):
        #     miniper = 0.24
        #     minidict[user] = miniper
        # elif (mini0 >= 209 and mini0 <= 245):
        #     miniper = 0.3
        #     minidict[user] = miniper

        # 鼠标离开
        # if (mini0 >= 31 and mini0 <= 73):
        #     miniper = 0.02
        #     minidict[user] = miniper
        # elif (mini0 >= 74 and mini0 <= 116):
        #     miniper = 0.04
        #     minidict[user] = miniper
        # elif (mini0 >= 117 and mini0 <= 159):
        #     miniper = 0.06
        #     minidict[user] = miniper
        # elif (mini0 >= 160 and mini0 <= 202):
        #     miniper = 0.08
        #     minidict[user] = miniper
        # elif (mini0 >= 203 and mini0 <= 245):
        #     miniper = 0.1
        #     minidict[user] = miniper

        # 最小化
        # if(mini0 >= 21 and mini0 <= 65):
        #     miniper = 0.02
        #     minidict[user] = miniper
        # elif (mini0 >= 66 and mini0 <= 110):
        #     miniper = 0.04
        #     minidict[user] = miniper
        # elif (mini0 >= 111 and mini0 <= 155):
        #     miniper = 0.06
        #     minidict[user] = miniper
        # elif (mini0 >= 156 and mini0 <= 200):
        #     miniper = 0.08
        #     minidict[user] = miniper
        # elif (mini0 >= 201 and mini0 <= 245):
        #     miniper = 0.1
        #     minidict[user] = miniper

print(len(minidict))

for row in range(2, workSheet.max_row+1):
    usernames = workSheet.cell(row, 1).value
    if usernames in minidict:
        workSheet.cell(row, 28, minidict[usernames])
        workBook.save(saveFile)
        print(row, minidict[usernames])
