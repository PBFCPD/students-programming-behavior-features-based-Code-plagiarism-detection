import openpyxl
from decimal import Decimal

cid = '2180'

data_path = r'../excel/2-' + cid + '/' + cid +'_total.xlsx'
workBook = openpyxl.load_workbook(data_path)
workSheet = workBook.worksheets[0]
saveFile = r'../excel/2-' + cid + '/' + cid +'_total.xlsx'

if(cid == '2232'):
    workSheet.cell(1, 13, '抄袭率abs')
else:
    workSheet.cell(1, 14, '抄袭率abs')

copy = 0
for row in range(2, workSheet.max_row+1):
    username = workSheet.cell(row, 1).value

    if(cid == '2232'):
        minisize = workSheet.cell(row, 3).value
        leave = workSheet.cell(row, 4).value
        command = workSheet.cell(row, 5).value
        abnarmal = workSheet.cell(row, 6).value
        timeabs = workSheet.cell(row, 7).value
        time2 = workSheet.cell(row, 8).value
        time3 = workSheet.cell(row, 9).value
        time4 = workSheet.cell(row, 10).value
        time5 = workSheet.cell(row, 11).value
    else:
        minisize = workSheet.cell(row, 3).value
        command = workSheet.cell(row, 4).value
        abnarmal = workSheet.cell(row, 5).value
        timeabs = workSheet.cell(row, 6).value
        time2 = workSheet.cell(row, 7).value
        time3 = workSheet.cell(row, 8).value
        time4 = workSheet.cell(row, 9).value
        time5 = workSheet.cell(row, 10).value
    print(minisize, command, abnarmal, timeabs, time2, time3, time4, time4)
    if (cid == '2232'):
        copyabs = Decimal(str(0.2 * minisize)) + Decimal(str(0.2 * leave)) + Decimal(str(0.1 * command)) + Decimal(str(0.3 * abnarmal)) + Decimal(
            str(0.2 * timeabs))
        copyabs = round(copyabs, 2)
        copy2 = Decimal(str(0.2 * minisize)) + Decimal(str(0.2 * leave)) + Decimal(str(0.1 * command)) + Decimal(str(0.3 * abnarmal)) + Decimal(
            str(0.2 * time2))
        copy2 = round(copy2, 2)
        copy3 = Decimal(str(0.2 * minisize)) + Decimal(str(0.2 * leave)) + Decimal(str(0.1 * command)) + Decimal(str(0.3 * abnarmal)) + Decimal(
            str(0.2 * time3))
        copy3 = round(copy3, 2)
        copy4 = Decimal(str(0.2 * minisize)) + Decimal(str(0.2 * leave)) + Decimal(str(0.1 * command)) + Decimal(str(0.3 * abnarmal)) + Decimal(
            str(0.2 * time4))
        copy4 = round(copy4, 2)
        copy5 = Decimal(str(0.2 * minisize)) + Decimal(str(0.2 * leave)) + Decimal(str(0.1 * command)) + Decimal(str(0.3 * abnarmal)) + Decimal(
            str(0.2 * time5))
        copy5 = round(copy5, 2)
    else:
        copyabs = Decimal(str(0.3 * minisize)) + Decimal(str(0.3 * command)) + Decimal(str(0.1 * abnarmal)) +Decimal(
            str(0.3 * timeabs))
        copyabs = round(copyabs, 2)
        copy2 = Decimal(str(0.3 * minisize)) + Decimal(str(0.3 * command)) + Decimal(str(0.1 * abnarmal)) + Decimal(
            str(0.3 * time2))
        copy2 = round(copy2, 2)
        copy3 = Decimal(str(0.3 * minisize)) + Decimal(str(0.3 * command)) + Decimal(str(0.1 * abnarmal)) + Decimal(
            str(0.3 * time3))
        copy3 = round(copy3, 2)
        copy4 = Decimal(str(0.3 * minisize)) + Decimal(str(0.3 * command)) + Decimal(str(0.1 * abnarmal)) + Decimal(
            str(0.3 * time4))
        copy4 = round(copy4, 2)
        copy5 = Decimal(str(0.3 * minisize)) + Decimal(str(0.3 * command)) + Decimal(str(0.1 * abnarmal)) + Decimal(
            str(0.3 * time5))
        copy5 = round(copy5, 2)

    print(copyabs, copy2, copy3, copy4, copy5)

    if (cid == '2232'):
        workSheet.cell(row, 13, copyabs)
        workSheet.cell(row, 14, copy2)
        workSheet.cell(row, 15, copy3)
        workSheet.cell(row, 16, copy4)
        workSheet.cell(row, 17, copy5)
    else:
        workSheet.cell(row, 17, copyabs)
        workSheet.cell(row, 18, copy2)
        workSheet.cell(row, 19, copy3)
        workSheet.cell(row, 20, copy4)
        workSheet.cell(row, 21, copy5)
    workBook.save(saveFile)
