import numpy as np
import pandas as pd
import openpyxl
from sklearn.metrics import accuracy_score
from sklearn.metrics import precision_score
from sklearn.metrics import recall_score
from sklearn.metrics import f1_score
from collections import Counter

data_path = r'../excel/res.xlsx'
data_frame = pd.read_excel(data_path, sheet_name='2315')
true8 = list(data_frame['true8'])
pre8 = list(data_frame['pre8'])
true85 = list(data_frame['true85'])
pre85 = list(data_frame['pre85'])
true9 = list(data_frame['true9'])
pre9 = list(data_frame['pre9'])
true95 = list(data_frame['true95'])
pre95 = list(data_frame['pre95'])
print("*****精确率*****")
print(precision_score(true8, pre8, average='weighted'))
print(precision_score(true85, pre85, average='weighted'))
print(precision_score(true9, pre9, average='weighted'))
print(precision_score(true95, pre95, average='weighted'))
print("*****召回率*****")
print(recall_score(true8, pre8, average='weighted'))
print(recall_score(true85, pre85, average='weighted'))
print(recall_score(true9, pre9, average='weighted'))
print(recall_score(true95, pre95, average='weighted'))
print("*****F1-score*****")
print(f1_score(true8, pre8, average='weighted'))
print(f1_score(true85, pre85, average='weighted'))
print(f1_score(true9, pre9, average='weighted'))
print(f1_score(true95, pre95, average='weighted'))

# resE = Counter(E)
# TN_E = resE['TN']
# FN_E = resE['FN']
# FP_E = resE['FP']
# TP_E = resE['TP']
# print(resE)
# precision = TP_E / (TP_E + FP_E)
# recall = TP_E / (TP_E + FN_E)
# accuracy = (TP_E + TN_E) / (TP_E + FP_E + TN_E + FN_E)
# F1_Score = 2*precision*recall/(precision+recall)
# print(precision, recall, accuracy, F1_Score)
#
# resEF = Counter(EF)
# TN_EF = resEF['TN']
# FN_EF = resEF['FN']
# FP_EF = resEF['FP']
# TP_EF = resEF['TP']
# print(resEF)
# precision = TP_EF / (TP_EF + FP_EF)
# recall = TP_EF / (TP_EF + FN_EF)
# accuracy = (TP_EF + TN_EF) / (TP_EF + FP_EF + TN_EF + FN_EF)
# F1_Score = 2*precision*recall/(precision+recall)
# print(precision, recall, accuracy, F1_Score)
#
# resN = Counter(N)
# TN_N = resN['TN']
# FN_N = resN['FN']
# FP_N = resN['FP']
# TP_N = resN['TP']
# print(resN)
# precision = TP_N / (TP_N + FP_N)
# recall = TP_N / (TP_N + FN_N)
# accuracy = (TP_N + TN_N) / (TP_N + FP_N + TN_N + FN_N)
# F1_Score = 2*precision*recall/(precision+recall)
# print(precision, recall, accuracy, F1_Score)
#
# resNF = Counter(NF)
# TN_NF = resNF['TN']
# FN_NF = resNF['FN']
# FP_NF = 0
# TP_NF = 0
# # FP_NF = resNF['FP']
# # TP_NF = resNF['TP']
# print(resNF)
# precision = TP_NF / (TP_NF + FP_NF)
# recall = TP_NF / (TP_NF + FN_NF)
# accuracy = (TP_NF + TN_NF) / (TP_NF + FP_NF + TN_NF + FN_NF)
# F1_Score = 2*precision*recall/(precision+recall)
# print(precision, recall, accuracy, F1_Score)

# data_path = r'../excel/res.xlsx'
# workBook = openpyxl.load_workbook(data_path)
# workSheet = workBook.worksheets[1]
# saveFile = r'../excel/res.xlsx'
#
# for row in range(2, workSheet.max_row+1):
#     true8 = workSheet.cell(row, 2).value
#     pre8 = workSheet.cell(row, 3).value
#
#     true85 = workSheet.cell(row, 6).value
#     pre85 = workSheet.cell(row, 7).value
#
#     true9 = workSheet.cell(row, 10).value
#     pre9 = workSheet.cell(row, 11).value
#
#     true95 = workSheet.cell(row, 14).value
#     pre95 = workSheet.cell(row, 15).value
#
#     if(true8 == pre8):
#         if(true8 == 'A'):
#             workSheet.cell(row, 4, 'TP')
#         elif(true8 == 'B'):
#             workSheet.cell(row, 4, 'TN')
#     else:
#         if (true8 == 'A'):
#             workSheet.cell(row, 4, 'FN')
#         elif (true8 == 'B'):
#             workSheet.cell(row, 4, 'FP')
#
#     if (true85 == pre85):
#         if (true85 == 'A'):
#             workSheet.cell(row, 8, 'TP')
#         elif (true85 == 'B'):
#             workSheet.cell(row, 8, 'TN')
#     else:
#         if (true85 == 'A'):
#             workSheet.cell(row, 8, 'FN')
#         elif (true85 == 'B'):
#             workSheet.cell(row, 8, 'FP')
#
#     if (true9 == pre9):
#         if (true9 == 'A'):
#             workSheet.cell(row, 12, 'TP')
#         elif (true9 == 'B'):
#             workSheet.cell(row, 12, 'TN')
#     else:
#         if (true9 == 'A'):
#             workSheet.cell(row, 12, 'FN')
#         elif (true9 == 'B'):
#             workSheet.cell(row, 12, 'FP')
#
#     if (true95 == pre95):
#         if (true95 == 'A'):
#             workSheet.cell(row, 16, 'TP')
#         elif (true95 == 'B'):
#             workSheet.cell(row, 16, 'TN')
#     else:
#         if (true95 == 'A'):
#             workSheet.cell(row, 16, 'FN')
#         elif (true95 == 'B'):
#             workSheet.cell(row, 16, 'FP')
#
#     workBook.save(saveFile)

