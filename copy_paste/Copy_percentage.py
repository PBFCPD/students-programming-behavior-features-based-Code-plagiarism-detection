import openpyxl
from decimal import Decimal

cid = '2262'

data_path = r'../excel/6-' + cid + '/' + cid +'_total.xlsx'
workBook = openpyxl.load_workbook(data_path)
workSheet = workBook.worksheets[0]
saveFile = r'../excel/6-' + cid + '/' + cid +'_total_3.xlsx'

# 17   22
c_save = 17

if(cid == '2232'):
    workSheet.cell(1, 13, '抄袭率abs')
    workSheet.cell(1, 14, '抄袭率0.2')
    workSheet.cell(1, 15, '抄袭率0.3')
    workSheet.cell(1, 16, '抄袭率0.4')
    workSheet.cell(1, 17, '抄袭率0.5')
elif(cid == '2262'):
    workSheet.cell(1, c_save, '抄袭率abs')
    workSheet.cell(1, c_save + 1, '抄袭率0.2')
    workSheet.cell(1, c_save + 2, '抄袭率0.3')
    workSheet.cell(1, c_save + 3, '抄袭率0.4')
    workSheet.cell(1, c_save + 4, '抄袭率0.5')
else:
    workSheet.cell(1, 12, '抄袭率abs')
    workSheet.cell(1, 13, '抄袭率0.2')
    workSheet.cell(1, 14, '抄袭率0.3')
    workSheet.cell(1, 15, '抄袭率0.4')
    workSheet.cell(1, 16, '抄袭率0.5')

copy = 0
for row in range(2, workSheet.max_row+1):
    username = workSheet.cell(row, 1).value

    if(cid == '2232'):
        minisize = workSheet.cell(row, 3).value
        leave = workSheet.cell(row, 4).value
        command = workSheet.cell(row, 5).value
        abnarmal = workSheet.cell(row, 6).value
        timeabs = workSheet.cell(row, 7).value
        time2 = workSheet.cell(row, 8).value
        time3 = workSheet.cell(row, 9).value
        time4 = workSheet.cell(row, 10).value
        time5 = workSheet.cell(row, 11).value
    elif(cid == '2262'):
        minisize = workSheet.cell(row, 3).value
        leaveweb = workSheet.cell(row, 4).value
        abnormal = workSheet.cell(row, 5).value
        commandabs = workSheet.cell(row, 6).value
        command2 = workSheet.cell(row, 7).value
        command3 = workSheet.cell(row, 8).value
        command4 = workSheet.cell(row, 9).value
        command5 = workSheet.cell(row, 10).value
        timeabs = workSheet.cell(row, 11).value
        time2 = workSheet.cell(row, 12).value
        time3 = workSheet.cell(row, 13).value
        time4 = workSheet.cell(row, 14).value
        time5 = workSheet.cell(row, 15).value
    else:
        minisize = workSheet.cell(row, 3).value
        command = workSheet.cell(row, 4).value
        abnarmal = workSheet.cell(row, 5).value
        timeabs = workSheet.cell(row, 6).value
        time2 = workSheet.cell(row, 7).value
        time3 = workSheet.cell(row, 8).value
        time4 = workSheet.cell(row, 9).value
        time5 = workSheet.cell(row, 10).value
    # print(minisize, command, abnarmal, timeabs, time2, time3, time4, time4)
    if (cid == '2232'):
        copyabs = Decimal(str(0.2 * minisize)) + Decimal(str(0.2 * leave)) + Decimal(str(0.1 * command)) + Decimal(str(0.3 * abnarmal)) + Decimal(
            str(0.2 * timeabs))
        copyabs = round(copyabs, 2)
        copy2 = Decimal(str(0.2 * minisize)) + Decimal(str(0.2 * leave)) + Decimal(str(0.1 * command)) + Decimal(str(0.3 * abnarmal)) + Decimal(
            str(0.2 * time2))
        copy2 = round(copy2, 2)
        copy3 = Decimal(str(0.2 * minisize)) + Decimal(str(0.2 * leave)) + Decimal(str(0.1 * command)) + Decimal(str(0.3 * abnarmal)) + Decimal(
            str(0.2 * time3))
        copy3 = round(copy3, 2)
        copy4 = Decimal(str(0.2 * minisize)) + Decimal(str(0.2 * leave)) + Decimal(str(0.1 * command)) + Decimal(str(0.3 * abnarmal)) + Decimal(
            str(0.2 * time4))
        copy4 = round(copy4, 2)
        copy5 = Decimal(str(0.2 * minisize)) + Decimal(str(0.2 * leave)) + Decimal(str(0.1 * command)) + Decimal(str(0.3 * abnarmal)) + Decimal(
            str(0.2 * time5))
        copy5 = round(copy5, 2)
    elif (cid == '2262'):
        # 0.15  0.15  0.3  0.2  0.2
        # 0.1  0.1  0.3  0.2  0.3
        c_mini = 0.2
        c_leave = 0.1
        c_abnormal = 0.2
        c_command = 0.2
        c_time = 0.3

        copyabs = Decimal(str(c_mini * minisize)) + Decimal(str(c_leave * leaveweb)) + Decimal(str(c_command * commandabs)) + Decimal(str(c_abnormal * abnormal)) + Decimal(
            str(c_time * timeabs))
        copyabs = round(copyabs, 2)

        copy2 = Decimal(str(c_mini * minisize)) + Decimal(str(c_leave * leaveweb)) + Decimal(str(c_command * command2)) + Decimal(str(c_abnormal * abnormal)) + Decimal(
            str(c_time * time2))
        copy2 = round(copy2, 2)

        copy3 = Decimal(str(c_mini * minisize)) + Decimal(str(c_leave * leaveweb)) + Decimal(str(c_command * command3)) + Decimal(str(c_abnormal * abnormal)) + Decimal(
            str(c_time * time3))
        copy3 = round(copy3, 2)

        copy4 = Decimal(str(c_mini * minisize)) + Decimal(str(c_leave * leaveweb)) + Decimal(str(c_command * command4)) + Decimal(str(c_abnormal * abnormal)) + Decimal(
            str(c_time * time4))
        copy4 = round(copy4, 2)

        copy5 = Decimal(str(c_mini * minisize)) + Decimal(str(c_leave * leaveweb)) + Decimal(str(c_command * command5)) + Decimal(str(c_abnormal * abnormal)) + Decimal(
            str(c_time * time5))
        copy5 = round(copy5, 2)
    else:
        copyabs = Decimal(str(0.3 * minisize)) + Decimal(str(0.3 * command)) + Decimal(str(0.1 * abnarmal)) +Decimal(
            str(0.3 * timeabs))
        copyabs = round(copyabs, 2)
        copy2 = Decimal(str(0.3 * minisize)) + Decimal(str(0.3 * command)) + Decimal(str(0.1 * abnarmal)) + Decimal(
            str(0.3 * time2))
        copy2 = round(copy2, 2)
        copy3 = Decimal(str(0.3 * minisize)) + Decimal(str(0.3 * command)) + Decimal(str(0.1 * abnarmal)) + Decimal(
            str(0.3 * time3))
        copy3 = round(copy3, 2)
        copy4 = Decimal(str(0.3 * minisize)) + Decimal(str(0.3 * command)) + Decimal(str(0.1 * abnarmal)) + Decimal(
            str(0.3 * time4))
        copy4 = round(copy4, 2)
        copy5 = Decimal(str(0.3 * minisize)) + Decimal(str(0.3 * command)) + Decimal(str(0.1 * abnarmal)) + Decimal(
            str(0.3 * time5))
        copy5 = round(copy5, 2)

    print(copyabs, copy2, copy3, copy4, copy5)

    if (cid == '2232'):
        workSheet.cell(row, 13, copyabs)
        workSheet.cell(row, 14, copy2)
        workSheet.cell(row, 15, copy3)
        workSheet.cell(row, 16, copy4)
        workSheet.cell(row, 17, copy5)
    elif (cid == '2262'):
        workSheet.cell(row, c_save, copyabs)
        workSheet.cell(row, c_save + 1, copy2)
        workSheet.cell(row, c_save + 2, copy3)
        workSheet.cell(row, c_save + 3, copy4)
        workSheet.cell(row, c_save + 4, copy5)
    else:
        workSheet.cell(row, 17, copyabs)
        workSheet.cell(row, 18, copy2)
        workSheet.cell(row, 19, copy3)
        workSheet.cell(row, 20, copy4)
        workSheet.cell(row, 21, copy5)
    workBook.save(saveFile)
