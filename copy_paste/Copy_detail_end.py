import openpyxl
import pandas as pd
from decimal import Decimal
from pandas import DataFrame,read_excel

cid = '2315'

data_path1 = r'../excel/7-' + cid + '/' + cid +'_detail.xlsx'
workBook1 = openpyxl.load_workbook(data_path1)
workSheet1 = workBook1.worksheets[0]

saveFile = r'../excel/7-' + cid + '/' + cid +'_detail_end.xlsx'

if(cid == '2232'):
    resultnum1 = pd.DataFrame(columns=["用户名", "竞赛", "最小化页面次数归一化", "鼠标离开页面的时间归一化", "终端输入次数归一化"])
elif (cid == '2315'):
    resultnum1 = pd.DataFrame(columns=["用户名", "竞赛", "最小化页面次数归一化", "鼠标离开页面的时间归一化", "终端输入次数abs归一化", "终端输入次数0.2归一化", "终端输入次数0.3归一化", "终端输入次数0.4归一化", "终端输入次数0.5归一化"])
else:
    resultnum1 = pd.DataFrame(columns=["用户名", "竞赛", "最小化页面次数归一化", "终端输入次数归一化"])

data = DataFrame(read_excel(data_path1))

minnisize = list(data['最小化页面次数'])

if(cid != '2315'):
    command = list(data['终端输入次数'])
    maxcommand = max(command)

maxminnisize = max(minnisize)


if(cid == '2232' or cid == '2262' or cid == '2315'):
    leave = list(data['鼠标离开页面的时间'])
    maxleave = max(leave)

# print(maxminnisize, maxcommand)

for row in range(2, workSheet1.max_row+1):
    username = workSheet1.cell(row, 1).value
    contest_id = workSheet1.cell(row, 2).value
    commandabs = workSheet1.cell(row, 22).value
    command2 = workSheet1.cell(row, 23).value
    command3 = workSheet1.cell(row, 24).value
    command4 = workSheet1.cell(row, 25).value
    command5 = workSheet1.cell(row, 26).value

    datamini = round(workSheet1.cell(row, 3).value / maxminnisize, 2)

    if(cid != '2315'):
        datacommand = round(workSheet1.cell(row, 6).value / maxcommand, 2)

    if(cid == '2232' or cid == '2262' or cid == '2315'):
        dataleave = round(workSheet1.cell(row, 4).value / maxleave, 2)

    if(cid == '2232'):
        resultnum1 = resultnum1.append(
            {"用户名": username, "竞赛": contest_id, "最小化页面次数归一化": datamini, "鼠标离开页面的时间归一化": dataleave,
             "终端输入次数归一化": datacommand}, ignore_index=True)
    elif (cid == '2315'):
        resultnum1 = resultnum1.append(
            {"用户名": username, "竞赛": contest_id, "最小化页面次数归一化": datamini, "鼠标离开页面的时间归一化": dataleave,
             "终端输入次数abs归一化": commandabs, "终端输入次数0.2归一化": command2, "终端输入次数0.3归一化": command3,
             "终端输入次数0.4归一化": command4, "终端输入次数0.5归一化": command5}, ignore_index=True)
    else:
        resultnum1 = resultnum1.append(
            {"用户名": username, "竞赛": contest_id, "最小化页面次数归一化": datamini, "终端输入次数归一化": datacommand}, ignore_index=True)


    resultnum1.to_excel(saveFile, index=False)





