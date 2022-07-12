import numpy as np
import pandas as pd
import openpyxl
from sklearn.metrics import accuracy_score
from sklearn.metrics import precision_score
from sklearn.metrics import recall_score
from sklearn.metrics import f1_score
from collections import Counter


data_path = r'../excel/6-2262/2262_total_2_消融实验.xlsx'
workBook = openpyxl.load_workbook(data_path)
workSheet = workBook.worksheets[4]
saveFile = r'../excel/6-2262/2262_total_2_消融实验.xlsx'

for row in range(2, workSheet.max_row):
    label = workSheet.cell(row, 2).value
    pre1 = workSheet.cell(row, 5).value
    pre2 = workSheet.cell(row, 9).value
    pre3 = workSheet.cell(row, 13).value
    pre4 = workSheet.cell(row, 17).value
    pre5 = workSheet.cell(row, 21).value
    # pre6 = workSheet.cell(row, 25).value
    # pre7 = workSheet.cell(row, 29).value
    # pre8 = workSheet.cell(row, 33).value
    # pre9 = workSheet.cell(row, 37).value
    # pre10 = workSheet.cell(row, 41).value

    print(label, pre5, pre4, pre3, pre2, pre1)
    if(label == pre1):
        if(label == 'A'):
            workSheet.cell(row, 7, 'TP')
        elif(label == 'B'):
            workSheet.cell(row, 7, 'TN')
    else:
        if (label == 'A'):
            workSheet.cell(row, 7, 'FN')
        elif (label == 'B'):
            workSheet.cell(row, 7, 'FP')

    if (label == pre2):
        if (label == 'A'):
            workSheet.cell(row, 11, 'TP')
        elif (label == 'B'):
            workSheet.cell(row, 11, 'TN')
    else:
        if (label == 'A'):
            workSheet.cell(row, 11, 'FN')
        elif (label == 'B'):
            workSheet.cell(row, 11, 'FP')

    if (label == pre3):
        if (label == 'A'):
            workSheet.cell(row, 15, 'TP')
        elif (label == 'B'):
            workSheet.cell(row, 15, 'TN')
    else:
        if (label == 'A'):
            workSheet.cell(row, 15, 'FN')
        elif (label == 'B'):
            workSheet.cell(row, 15, 'FP')

    if (label == pre4):
        if (label == 'A'):
            workSheet.cell(row, 19, 'TP')
        elif (label == 'B'):
            workSheet.cell(row, 19, 'TN')
    else:
        if (label == 'A'):
            workSheet.cell(row, 19, 'FN')
        elif (label == 'B'):
            workSheet.cell(row, 19, 'FP')

    if (label == pre5):
        if (label == 'A'):
            workSheet.cell(row, 23, 'TP')
        elif (label == 'B'):
            workSheet.cell(row, 23, 'TN')
    else:
        if (label == 'A'):
            workSheet.cell(row, 23, 'FN')
        elif (label == 'B'):
            workSheet.cell(row, 23, 'FP')

    # if (label == pre6):
    #     if (label == 'A'):
    #         workSheet.cell(row, 27, 'TP')
    #     elif (label == 'B'):
    #         workSheet.cell(row, 27, 'TN')
    # else:
    #     if (label == 'A'):
    #         workSheet.cell(row, 27, 'FN')
    #     elif (label == 'B'):
    #         workSheet.cell(row, 27, 'FP')
    #
    # if (label == pre7):
    #     if (label == 'A'):
    #         workSheet.cell(row, 31, 'TP')
    #     elif (label == 'B'):
    #         workSheet.cell(row, 31, 'TN')
    # else:
    #     if (label == 'A'):
    #         workSheet.cell(row, 31, 'FN')
    #     elif (label == 'B'):
    #         workSheet.cell(row, 31, 'FP')
    #
    # if (label == pre8):
    #     if (label == 'A'):
    #         workSheet.cell(row, 35, 'TP')
    #     elif (label == 'B'):
    #         workSheet.cell(row, 35, 'TN')
    # else:
    #     if (label == 'A'):
    #         workSheet.cell(row, 35, 'FN')
    #     elif (label == 'B'):
    #         workSheet.cell(row, 35, 'FP')
    #
    # if (label == pre9):
    #     if (label == 'A'):
    #         workSheet.cell(row, 39, 'TP')
    #     elif (label == 'B'):
    #         workSheet.cell(row, 39, 'TN')
    # else:
    #     if (label == 'A'):
    #         workSheet.cell(row, 39, 'FN')
    #     elif (label == 'B'):
    #         workSheet.cell(row, 39, 'FP')
    #
    # if (label == pre10):
    #     if (label == 'A'):
    #         workSheet.cell(row, 43, 'TP')
    #     elif (label == 'B'):
    #         workSheet.cell(row, 43, 'TN')
    # else:
    #     if (label == 'A'):
    #         workSheet.cell(row, 43, 'FN')
    #     elif (label == 'B'):
    #         workSheet.cell(row, 43, 'FP')
    print(row)
    workBook.save(saveFile)
