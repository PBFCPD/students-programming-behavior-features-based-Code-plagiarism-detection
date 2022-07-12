# 按照比例计算 2315

# 1.最小化页面次数归一化  -  0.1
# 2.鼠标离开页面的时间归一化  -  0.1
# 3.异常粘贴次数  -  0.3
# 4.终端输入次数abs归一化	终端输入次数0.2归一化	终端输入次数0.3归一化	终端输入次数0.4归一化	终端输入次数0.5归一化  -  0.2
# 5.timeabs	time0.2	time0.3	time0.4	time0.5  -  0.3

import openpyxl
import pandas as pd
from decimal import Decimal
from pandas import DataFrame, read_excel

c_mini = 0.1
c_leave = 0.1
c_abnormal = 0.3
c_command = 0.2
c_time = 0.3

cid = '2315'
data_path = r'../excel/7-' + cid + '/' + cid +'_total.xlsx'
workBook = openpyxl.load_workbook(data_path)
workSheet = workBook.worksheets[0]

saveFile = r'../excel/7-' + cid + '/' + cid +'_total.xlsx'


workSheet.cell(1, 17, '最小化页面次数归一化')
workSheet.cell(1, 18, '鼠标离开页面的时间归一化')
workSheet.cell(1, 19, '异常粘贴次数归一化')
workSheet.cell(1, 20, '终端输入次数abs归一化')
workSheet.cell(1, 21, '终端输入次数0.2归一化')
workSheet.cell(1, 22, '终端输入次数0.3归一化')
workSheet.cell(1, 23, '终端输入次数0.4归一化')
workSheet.cell(1, 24, '终端输入次数0.5归一化')
workSheet.cell(1, 25, 'timeabs')
workSheet.cell(1, 26, 'time0.2')
workSheet.cell(1, 27, 'time0.3')
workSheet.cell(1, 28, 'time0.4')
workSheet.cell(1, 29, 'time0.5')

data_path1 = r'../excel/7-' + cid + '/' + cid +'_total.xlsx'
data1 = DataFrame(read_excel(data_path1, sheet_name='Sheet1'))

datamini = data1.sort_values(by="time0.5", ascending=True)
username = list(datamini['用户名'])
minisize = list(datamini['time0.5'])
mini0 = 0
minidict = {}
for user, mini in zip(list(username), list(minisize)):
    if(mini == 0):
        mini0 = mini0 + 1
        miniper = 0
        minidict[user] = miniper
    else:
        mini0 = mini0 + 1
        # time
        if (mini0 >= 2 and mini0 <= 45):
            miniper = 0.06
            minidict[user] = miniper
        elif (mini0 >= 46 and mini0 <= 89):
            miniper = 0.12
            minidict[user] = miniper
        elif (mini0 >= 90 and mini0 <= 133):
            miniper = 0.18
            minidict[user] = miniper
        elif (mini0 >= 134 and mini0 <= 178):
            miniper = 0.24
            minidict[user] = miniper
        elif (mini0 >= 179 and mini0 <= 223):
            miniper = 0.3
            minidict[user] = miniper

        # 终端5
        # if (mini0 >= 15 and mini0 <= 55):
        #     miniper = 0.04
        #     minidict[user] = miniper
        # elif (mini0 >= 56 and mini0 <= 97):
        #     miniper = 0.08
        #     minidict[user] = miniper
        # elif (mini0 >= 98 and mini0 <= 139):
        #     miniper = 0.12
        #     minidict[user] = miniper
        # elif (mini0 >= 140 and mini0 <= 181):
        #     miniper = 0.16
        #     minidict[user] = miniper
        # elif (mini0 >= 182 and mini0 <= 223):
        #     miniper = 0.2
        #     minidict[user] = miniper

        # 终端4
        # if (mini0 >= 13 and mini0 <= 54):
        #     miniper = 0.04
        #     minidict[user] = miniper
        # elif (mini0 >= 55 and mini0 <= 96):
        #     miniper = 0.08
        #     minidict[user] = miniper
        # elif (mini0 >= 97 and mini0 <= 138):
        #     miniper = 0.12
        #     minidict[user] = miniper
        # elif (mini0 >= 139 and mini0 <= 180):
        #     miniper = 0.16
        #     minidict[user] = miniper
        # elif (mini0 >= 181 and mini0 <= 223):
        #     miniper = 0.2
        #     minidict[user] = miniper

        # 终端3
        # if (mini0 >= 35 and mini0 <= 71):
        #     miniper = 0.04
        #     minidict[user] = miniper
        # elif (mini0 >= 72 and mini0 <= 109):
        #     miniper = 0.08
        #     minidict[user] = miniper
        # elif (mini0 >= 110 and mini0 <= 147):
        #     miniper = 0.12
        #     minidict[user] = miniper
        # elif (mini0 >= 148 and mini0 <= 185):
        #     miniper = 0.16
        #     minidict[user] = miniper
        # elif (mini0 >= 186 and mini0 <= 223):
        #     miniper = 0.2
        #     minidict[user] = miniper

        # 终端2
        # if (mini0 >= 24 and mini0 <= 63):
        #     miniper = 0.04
        #     minidict[user] = miniper
        # elif (mini0 >= 64 and mini0 <= 103):
        #     miniper = 0.08
        #     minidict[user] = miniper
        # elif (mini0 >= 103 and mini0 <= 142):
        #     miniper = 0.12
        #     minidict[user] = miniper
        # elif (mini0 >= 143 and mini0 <= 182):
        #     miniper = 0.16
        #     minidict[user] = miniper
        # elif (mini0 >= 183 and mini0 <= 223):
        #     miniper = 0.2
        #     minidict[user] = miniper

        # 终端abs
        # if (mini0 >= 12 and mini0 <= 53):
        #     miniper = 0.04
        #     minidict[user] = miniper
        # elif (mini0 >= 54 and mini0 <= 95):
        #     miniper = 0.08
        #     minidict[user] = miniper
        # elif (mini0 >= 96 and mini0 <= 137):
        #     miniper = 0.12
        #     minidict[user] = miniper
        # elif (mini0 >= 138 and mini0 <= 180):
        #     miniper = 0.16
        #     minidict[user] = miniper
        # elif (mini0 >= 181 and mini0 <= 223):
        #     miniper = 0.2
        #     minidict[user] = miniper

        # 异常粘贴
        # if (mini0 >= 20 and mini0 <= 59):
        #     miniper = 0.06
        #     minidict[user] = miniper
        # elif (mini0 >= 60 and mini0 <= 100):
        #     miniper = 0.12
        #     minidict[user] = miniper
        # elif (mini0 >= 101 and mini0 <= 141):
        #     miniper = 0.18
        #     minidict[user] = miniper
        # elif (mini0 >= 142 and mini0 <= 182):
        #     miniper = 0.24
        #     minidict[user] = miniper
        # elif (mini0 >= 183 and mini0 <= 223):
        #     miniper = 0.3
        #     minidict[user] = miniper

        # 鼠标离开
        # if (mini0 >= 16 and mini0 <= 56):
        #     miniper = 0.02
        #     minidict[user] = miniper
        # elif (mini0 >= 57 and mini0 <= 97):
        #     miniper = 0.04
        #     minidict[user] = miniper
        # elif (mini0 >= 98 and mini0 <= 139):
        #     miniper = 0.06
        #     minidict[user] = miniper
        # elif (mini0 >= 140 and mini0 <= 181):
        #     miniper = 0.08
        #     minidict[user] = miniper
        # elif (mini0 >= 182 and mini0 <= 223):
        #     miniper = 0.1
        #     minidict[user] = miniper

        # 最小化
        # if(mini0 >= 9 and mini0 <= 51):
        #     miniper = 0.02
        #     minidict[user] = miniper
        # elif (mini0 >= 52 and mini0 <= 94):
        #     miniper = 0.04
        #     minidict[user] = miniper
        # elif (mini0 >= 95 and mini0 <= 137):
        #     miniper = 0.06
        #     minidict[user] = miniper
        # elif (mini0 >= 138 and mini0 <= 180):
        #     miniper = 0.08
        #     minidict[user] = miniper
        # elif (mini0 >= 181 and mini0 <= 223):
        #     miniper = 0.1
        #     minidict[user] = miniper

print(len(minidict))

for row in range(2, workSheet.max_row+1):
    usernames = workSheet.cell(row, 1).value
    if usernames in minidict:
        workSheet.cell(row, 29, minidict[usernames])
        workBook.save(saveFile)
        print(row, minidict[usernames])
