import openpyxl
import pandas as pd
from decimal import Decimal
from pandas import DataFrame,read_excel

cid = '2315'

data_path1 = r'../excel/7-' + cid + '/' + cid +'_time.xlsx'
workBook1 = openpyxl.load_workbook(data_path1)
workSheet1 = workBook1.worksheets[0]

saveFile = r'../excel/7-' + cid + '/' + cid +'_time_end.xlsx'

resultnum1 = pd.DataFrame(columns=["用户名", "竞赛", "abs", "0.2", "0.3", "0.4", "0.5"])

if (cid == '2167'):
    count = 14
elif(cid == '2315'):
    count = 13
else:
    count = 10
key = 0
print(workSheet1.max_row+1)
while(15 + key * count <= workSheet1.max_row+1):
    abs = 0
    beyondtime_2 = 0
    beyondtime_3 = 0
    beyondtime_4 = 0
    beyondtime_5 = 0
    first = 2 + key * count
    last = 2 + count + key * count
    for j in range(first, last):
        username = workSheet1.cell(j, 1).value
        contest_id = workSheet1.cell(j, 2).value
        abs = Decimal(str(abs))
        beyondtime_2 = Decimal(str(beyondtime_2))
        beyondtime_3 = Decimal(str(beyondtime_3))
        beyondtime_4 = Decimal(str(beyondtime_4))
        beyondtime_5 = Decimal(str(beyondtime_5))

        t1 = Decimal(str(workSheet1.cell(j, 21).value))
        t2 = Decimal(str(workSheet1.cell(j, 22).value))
        t3 = Decimal(str(workSheet1.cell(j, 23).value))
        t4 = Decimal(str(workSheet1.cell(j, 24).value))
        t5 = Decimal(str(workSheet1.cell(j, 25).value))

        abs = abs + t1
        beyondtime_2 = beyondtime_2 + t2
        beyondtime_3 = beyondtime_3 + t3
        beyondtime_4 = beyondtime_4 + t4
        beyondtime_5 = beyondtime_5 + t5
        print("原始", t1, t2, t3, t4, t5)
        print("求和", abs, beyondtime_2, beyondtime_3, beyondtime_4, beyondtime_5)
        print("***********************")
    key = key + 1
    abs = round(abs, 2)
    beyondtime_2 = round(beyondtime_2, 2)
    beyondtime_3 = round(beyondtime_3, 2)
    beyondtime_4 = round(beyondtime_4, 2)
    beyondtime_5 = round(beyondtime_5, 2)

    print(first, last, key, username, contest_id, abs, beyondtime_2, beyondtime_3, beyondtime_4, beyondtime_5)
    print("---------------------------------------------------------------")
    resultnum1 = resultnum1.append(
        {"用户名": username, "竞赛": contest_id, "abs": abs, "0.2": beyondtime_2,"0.3":beyondtime_3,"0.4":beyondtime_4,"0.5":beyondtime_5}, ignore_index=True)

    resultnum1.to_excel(saveFile, index=False)



