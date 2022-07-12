import MySQLdb
import pandas as pd
from datetime import datetime
import openpyxl

db1 = MySQLdb.connect("202.4.155.100", "root", "*", "*", charset='utf8')
# db2 = MySQLdb.connect("202.4.155.100", "root", "*", "*", charset='utf8')

cursor1 = db1.cursor()
# cursor2 = db2.cursor()

cid = '2262'

data_path1 = r'../excel/6-' + cid + '/' + cid +'_total_2.xlsx'
workBook1 = openpyxl.load_workbook(data_path1)
workSheet1 = workBook1.worksheets[2]
# print(workSheet1.max_row+1)
for row in range(2, 246):
    username = workSheet1.cell(row, 1).value
    contest_id = workSheet1.cell(row, 2).value
    avesimilar = workSheet1.cell(row, 16).value
    copy_percent = round(workSheet1.cell(row, 34).value, 2)
    label_3 = workSheet1.cell(row, 51).value
    our_3 = workSheet1.cell(row, 52).value
    label_2 = workSheet1.cell(row, 55).value
    our_2 = workSheet1.cell(row, 56).value

    sql = "INSERT INTO copy_detail (username,contest_id,avesimilar,copy_percent,label_3,our_3,label_2,our_2) " \
          "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)"
    cursor1.execute(sql, (username, contest_id, avesimilar, copy_percent, label_3, our_3, label_2, our_2))
    db1.commit()
    print(username, contest_id, avesimilar, copy_percent, label_3, our_3, label_2, our_2)


